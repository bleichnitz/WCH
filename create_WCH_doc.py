# ----- IMPORT LIBRARIES & CODE ----------------------------------------------------------------------------------------
# The following imports are from the PyPi repository.
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from colorama import Fore

# The following imports are from other *.py files within this project.
from load_files import load_file
from date_time import today_date

def create_doc(data):
    """
    This function takes the consolidated data that is now contained within a single array, and creates the functional
    "Who Can Help" document.
    :param data: the consolidated data array
    :return: n/a
    """
    workbook = load_file(file_name="Templates/WCH Template.xlsx")

    sheets = workbook.sheetnames
    write_sheet = workbook[sheets[0]]

    starting_row = 2
    row_counter = 0
    starting_col = 1
    col_counter = 0

    thick_border = Border(left=Side(style=None),
                          right=Side(style='thick'),
                          top=Side(style=None),
                          bottom=Side(style=None)
                          )

    thin_border = Border(left=Side(style=None),
                          right=Side(style='thin'),
                          top=Side(style=None),
                          bottom=Side(style=None)
                          )

    alt_row_colour = PatternFill("solid", start_color="c2c4c2")

    for row in data:
        #print(Fore.WHITE + f"{row}")

        for key in row:

            #print(Fore.CYAN + f"{key} / {row[key]}")
            current_row = starting_row + row_counter
            row_dimension = write_sheet.row_dimensions[current_row]
            row_dimension.height = 80
            current_col = starting_col + col_counter
            current_cell = write_sheet.cell(current_row, current_col)
            current_cell.value = row[key]
            current_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            if (current_row % 2) != 0:
                current_cell.fill = alt_row_colour
            if current_col == 3 or current_col == 8:
                current_cell.border = thick_border
            if current_col == 13:
                current_cell.border = thin_border

            col_counter += 1

        row_counter += 1
        col_counter = 0

    print(Fore.LIGHTYELLOW_EX + f"----- Who Can Help Document Created -----")
    date = today_date()
    file_name = f"{date} Who Can Help"
    workbook.save(f"{file_name}.xlsx")

